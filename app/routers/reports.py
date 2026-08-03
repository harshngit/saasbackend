from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import require_permission
from app.models import User
from app.services import report_service

router = APIRouter(prefix="/reports", tags=["reports"])

_view = require_permission("reports", "view")
_export = require_permission("reports", "export")


def _org_id(user: User) -> str:
    if not user.organization_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No organization on this account")
    return user.organization_id


def _to_xlsx(report: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([f"Report: {report['type']}"])
    if report.get("date_from") or report.get("date_to"):
        ws.append([f"Period: {report.get('date_from') or '—'} to {report.get('date_to') or '—'}"])
    ws.append([])
    ws.append(["Summary"])
    for key, value in report["summary"].items():
        ws.append([key, value])
    ws.append([])
    rows = report["rows"]
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_pdf(report: dict) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    pdf = FPDF(orientation="L")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Report: {report['type']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=9)
    if report.get("date_from") or report.get("date_to"):
        pdf.cell(0, 6, f"Period: {report.get('date_from') or '-'} to {report.get('date_to') or '-'}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for key, value in report["summary"].items():
        pdf.cell(0, 6, f"{key}: {value}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(3)
    rows = report["rows"]
    if rows:
        headers = list(rows[0].keys())
        col_w = max(20, min(60, 270 // max(1, len(headers))))
        pdf.set_font("Helvetica", "B", 8)
        for h in headers:
            pdf.cell(col_w, 6, str(h)[:22], border=1)
        pdf.ln()
        pdf.set_font("Helvetica", size=8)
        for r in rows[:500]:
            for h in headers:
                val = r.get(h)
                pdf.cell(col_w, 6, ("" if val is None else str(val))[:22], border=1)
            pdf.ln()
    return bytes(pdf.output())


@router.get("/{report_type}")
def get_report(
    report_type: str,
    user: User = Depends(_view),
    date_from: str | None = Query(default=None, description="YYYY-MM-DD"),
    date_to: str | None = Query(default=None, description="YYYY-MM-DD"),
    db: Session = Depends(get_db),
) -> dict:
    """A financial/transaction report — returns { summary, rows, type, date_from, date_to }."""
    return report_service.build_report(db, _org_id(user), report_type, date_from, date_to)


@router.get("/{report_type}/export")
def export_report(
    report_type: str,
    user: User = Depends(_export),
    format: str = Query(default="excel", description="excel | pdf"),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    db: Session = Depends(get_db),
) -> Response:
    """Download the report as an Excel (.xlsx) or PDF file."""
    report = report_service.build_report(db, _org_id(user), report_type, date_from, date_to)
    if format == "pdf":
        content = _to_pdf(report)
        media = "application/pdf"
        ext = "pdf"
    elif format == "excel":
        content = _to_xlsx(report)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="format must be 'excel' or 'pdf'")
    filename = f"{report_type}-report.{ext}"
    return Response(
        content=content,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
